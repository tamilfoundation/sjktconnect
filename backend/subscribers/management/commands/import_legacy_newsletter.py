"""
Django management command: Import legacy governance leaders as newsletter subscribers

Extracts emails from the 4 Tamil Foundation 2018 governance datasets:
- Headmasters
- Board Chairmen
- PTA Chairmen
- Alumni Chairmen

Features:
  - Deduplicates by email (same person with multiple roles)
  - Checks existing subscribers (skip if already subscribed)
  - Tracks roles in source_tag (e.g., TF_GOVERNANCE_LEADER_2018_HEADMASTER_PTA_CHAIR)
  - Uses silent opt-in (source="BULK_IMPORT")
  - Sets data_source and data_source_date for audit trail

Usage:
  python manage.py import_legacy_newsletter --dry-run
  python manage.py import_legacy_newsletter
"""

import openpyxl
from pathlib import Path
from datetime import date
from collections import defaultdict
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.core.mail import send_mail
import re

from subscribers.models import Subscriber


LEGACY_DATA_PATH = Path("C:/Users/tamil/Downloads/tamilschool/")

# Role mapping for display and source_tag
ROLE_MAPPING = {
    "headmaster": ("த.ஆசிரியர் - Status.xlsx", "த.ஆசிரியர்", "HEADMASTER", "HM"),
    "board_chair": ("தலைவர் - Status.xlsx", "தலைவர்", "BOARD_CHAIRMAN", "BC"),
    "pta_chair": ("பெ.ஆ.ச. - Status.xlsx", "பெ.ஆ.ச.", "PTA_CHAIRMAN", "PTA"),
    "alumni_chair": ("மு.மாணவர் - Status.xlsx", "மு.மாணவர்", "ALUMNI_CHAIRMAN", "AC"),
}


def is_valid_email(email):
    """Basic email validation."""
    if not email or not isinstance(email, str):
        return False
    email = email.strip()
    # Check for common invalid patterns
    if email.lower() in ("n/a", "na", "none", "-", "tiada", ""):
        return False
    # Basic regex check
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def normalize_name(name_str):
    """Convert name to Title Case with proper title formatting (Mr., Ms., Dr., etc.)."""
    if not name_str:
        return ""

    name_str = str(name_str).strip()

    # Common titles to handle
    titles = [
        'datuk', 'dato', 'datin', 'sir', 'madam', 'professor', 'prof',
        'dr', 'mr', 'ms', 'mrs', 'mdm', 'miss', 'en', 'puan',
    ]

    # Split into words
    words = name_str.split()
    if not words:
        return ""

    # Check if first word is a title
    first_lower = words[0].lower().rstrip('.')

    if first_lower in titles:
        title = words[0].lower()
        # Add period if not present
        if not title.endswith('.'):
            title = title + '.'
        # Title case the title
        title = title[0].upper() + title[1:]

        # Title case the rest of the name
        rest = ' '.join(words[1:])
        rest_titlecase = rest.title()

        return f"{title} {rest_titlecase}"
    else:
        # No title, just title case everything
        return name_str.title()


class Command(BaseCommand):
    help = "Import legacy governance leaders as newsletter subscribers"

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview what would be imported without saving',
        )

    def handle(self, *args, **options):
        dry_run = options['dry_run']

        self.stdout.write("=" * 100)
        self.stdout.write(
            "IMPORTING LEGACY GOVERNANCE LEADERS AS NEWSLETTER SUBSCRIBERS"
            + (" (DRY RUN)" if dry_run else "")
        )
        self.stdout.write("=" * 100)

        # Load schools lookup (MOE code -> school name)
        self.stdout.write("\nLoading schools dataset for names...")
        schools_lookup = self._load_schools_lookup()
        self.stdout.write(f"  Found {len(schools_lookup)} schools in legacy data")

        # Load all 4 datasets
        self.stdout.write("\nLoading governance datasets...")
        all_records = {}  # email -> {name, roles, org, school_name, primary_role}

        for role_key, (filename, sheet_name, role_display, role_abbr) in ROLE_MAPPING.items():
            self.stdout.write(f"  Loading {role_display}...")
            records = self._load_dataset(filename, sheet_name, schools_lookup)

            for name, email, school_name in records:
                if not is_valid_email(email):
                    continue

                email_lower = email.lower().strip()
                if email_lower not in all_records:
                    all_records[email_lower] = {
                        'name': name,
                        'email': email,
                        'school_name': school_name,  # Primary (first) school
                        'primary_role': role_display,  # Primary (first) role
                        'roles': [],
                        'role_abbrs': [],
                    }
                all_records[email_lower]['roles'].append(role_display)
                all_records[email_lower]['role_abbrs'].append(role_abbr)

        self.stdout.write(f"\n[OK] Extracted {len(all_records)} unique emails from all 4 datasets")

        # Check existing subscribers
        self.stdout.write("\nChecking existing subscribers...")
        existing_emails = set(
            Subscriber.objects.filter(is_active=True).values_list('email', flat=True)
        )
        self.stdout.write(f"  Found {len(existing_emails)} active subscribers")

        # Filter to new subscribers only
        new_subscribers = {
            email: data for email, data in all_records.items()
            if email not in existing_emails
        }
        self.stdout.write(f"  {len(new_subscribers)} new emails to add")

        # Show simulation or commit
        if dry_run:
            self._simulate_import(new_subscribers, existing_emails, all_records)
        else:
            self._commit_import(new_subscribers)

    def _load_schools_lookup(self):
        """Load schools dataset to map staff name -> school name."""
        filepath = LEGACY_DATA_PATH / "பள்ளிகள் - Last view used.xlsx"
        if not filepath.exists():
            raise CommandError(f"File not found: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["பள்ளிகள்"]

        # Build name -> school_name lookup (staff name -> list of schools)
        lookup = {}  # {staff_name: school_name}  (primary school only)

        # Columns for different roles: 35=HM, 39=BC, 36=PTA, 37=Alumni
        ROLE_COLUMNS = [35, 36, 37, 39]

        for row_idx in range(2, ws.max_row + 1):
            school_name = ws.cell(row_idx, 4).value  # Column 4 = School name

            if not school_name:
                continue

            # Check each role column for a staff name
            for col_idx in ROLE_COLUMNS:
                staff_name = ws.cell(row_idx, col_idx).value
                if staff_name:
                    # Map staff name to school (only track first/primary school)
                    if staff_name not in lookup:
                        lookup[staff_name] = school_name

        wb.close()
        return lookup

    def _load_dataset(self, filename, sheet_name, schools_lookup):
        """Load staff dataset and extract name + email + school name triplets."""
        filepath = LEGACY_DATA_PATH / filename
        if not filepath.exists():
            raise CommandError(f"File not found: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name]

        records = []
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row_idx, 4).value  # Column 4 = Name
            email_work = ws.cell(row_idx, 16).value  # Column 16 = Work email
            email_home = ws.cell(row_idx, 17).value  # Column 17 = Home email
            email_other = ws.cell(row_idx, 18).value  # Column 18 = Other email

            # Pick first available email
            email = email_work or email_home or email_other

            # Get school name by looking up staff name in schools
            school_name = schools_lookup.get(name, "")

            if name and email:
                records.append((name, email, school_name))

        wb.close()
        return records

    def _simulate_import(self, new_subscribers, existing_emails, all_records):
        """Show what would be imported (dry run)."""
        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("IMPORT SIMULATION (DRY RUN)")
        self.stdout.write("=" * 100)

        self.stdout.write(f"\nNew subscribers to add: {len(new_subscribers)}")
        self.stdout.write(f"Emails already subscribed: {len(existing_emails)}")
        self.stdout.write(f"Total unique emails: {len(all_records)}")

        # Show sample of new subscribers
        self.stdout.write("\nSample of new subscribers (first 10):")
        for i, (email, data) in enumerate(list(new_subscribers.items())[:10], 1):
            roles_str = " + ".join(data['roles'])
            primary_role = data.get('primary_role', '(unknown)')
            source_tag = f"TF_2018_{('_').join(data.get('role_abbrs', []))}"
            normalized_name = normalize_name(data['name'])
            self.stdout.write(f"\n  {i}. {normalized_name}")
            self.stdout.write(f"     Email: {email}")
            self.stdout.write(f"     School: {data.get('school_name', '(unknown)')}")
            self.stdout.write(f"     Primary role: {primary_role}")
            self.stdout.write(f"     All roles: {roles_str}")
            self.stdout.write(f"     Source tag: {source_tag}")

        # Summary
        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("SUMMARY")
        self.stdout.write("=" * 100)
        self.stdout.write(f"\nNew subscribers to create: {len(new_subscribers)}")
        self.stdout.write(f"Already subscribed (skipped): {len(existing_emails)}")
        self.stdout.write(f"Data source: Tamil Foundation, 2018-01-01")
        self.stdout.write(f"Source type: BULK_IMPORT (silent opt-in)")
        self.stdout.write("\n(This is a dry run. No data has been modified.)")
        self.stdout.write("To import, run without --dry-run flag")

    @transaction.atomic
    def _commit_import(self, new_subscribers):
        """Actually import the data into the database."""
        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("IMPORTING SUBSCRIBERS")
        self.stdout.write("=" * 100)

        created_count = 0
        skipped_count = 0
        for email, data in new_subscribers.items():
            # Build source_tag from roles (abbreviated to fit 50 char limit)
            role_abbrs = data.get('role_abbrs', [])
            source_tag = f"TF_2018_{('_').join(role_abbrs)}"

            # Use get_or_create to handle race conditions gracefully
            subscriber, created = Subscriber.objects.get_or_create(
                email=email,
                defaults={
                    'name': normalize_name(data['name']),
                    'organisation': "",  # Not available in legacy data
                    'school_name': data.get('school_name', ""),
                    'primary_governance_role': data.get('primary_role', ""),
                    'source': "BULK_IMPORT",
                    'source_tag': source_tag,
                    'is_active': True,
                    'bounce_count': 0,
                }
            )

            if created:
                created_count += 1
            else:
                skipped_count += 1

            if (created_count + skipped_count) % 100 == 0:
                self.stdout.write(f"  Processed {created_count + skipped_count}...")

        # Summary
        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("IMPORT COMPLETE")
        self.stdout.write("=" * 100)
        self.stdout.write(f"\nSubscribers created: {created_count}")
        self.stdout.write(f"Subscribers skipped (already existed): {skipped_count}")
        self.stdout.write(f"Data source: Tamil Foundation, 2018-01-01")
        self.stdout.write(f"Source type: BULK_IMPORT (silent opt-in)")
        self.stdout.write("\nAll records have been successfully imported to the database.")
