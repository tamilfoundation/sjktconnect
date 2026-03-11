"""
Import subscribers from an Excel file (donors/members list).

Reads name, email, status, and tag columns. Creates Subscriber records
with all 3 preferences enabled. Skips duplicates (existing emails).

Usage:
    python manage.py import_subscribers path/to/file.xlsx
    python manage.py import_subscribers path/to/file.xlsx --dry-run
"""

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

from subscribers.models import Subscriber, SubscriptionPreference


# Tamil column headers → field mapping
COLUMN_MAP = {
    "பெயர்": "name",
    "நிலை": "status",
    "மின்னஞ்சல் - Home": "email",
    "Tag": "tag",
}


class Command(BaseCommand):
    help = "Import subscribers from a donors/members Excel file."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to the Excel file")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Preview what would be imported without saving",
        )

    def handle(self, *args, **options):
        filepath = options["file"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
        except Exception as e:
            raise CommandError(f"Cannot open file: {e}")

        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise CommandError("File is empty.")

        # Map headers
        headers = rows[0]
        col_index = {}
        for i, header in enumerate(headers):
            if header in COLUMN_MAP:
                col_index[COLUMN_MAP[header]] = i

        required = {"name", "email"}
        missing = required - set(col_index.keys())
        if missing:
            raise CommandError(f"Missing columns: {missing}")

        created = 0
        skipped = 0
        invalid = 0

        for row in rows[1:]:
            email = row[col_index["email"]]
            if not email or not str(email).strip():
                invalid += 1
                continue

            email = str(email).strip().lower()

            try:
                validate_email(email)
            except ValidationError:
                self.stderr.write(f"  Invalid email: {email}")
                invalid += 1
                continue

            name = str(row[col_index["name"]] or "").strip()
            status = str(row[col_index.get("status", "")] or "").strip() if "status" in col_index else ""
            tag = str(row[col_index.get("tag", "")] or "").strip() if "tag" in col_index else ""

            if dry_run:
                self.stdout.write(f"  {email} — {name} [{status}] ({tag})")
                created += 1
                continue

            if Subscriber.objects.filter(email=email).exists():
                skipped += 1
                continue

            subscriber = Subscriber.objects.create(
                email=email,
                name=name,
                source="BULK_IMPORT",
                source_tag=tag,
                donor_status=status,
                is_active=True,
            )

            # Create all 3 preferences enabled
            for category, _ in SubscriptionPreference.CATEGORY_CHOICES:
                SubscriptionPreference.objects.create(
                    subscriber=subscriber,
                    category=category,
                    is_enabled=True,
                )

            created += 1

        prefix = "DRY RUN — would create" if dry_run else "Created"
        self.stdout.write(
            self.style.SUCCESS(
                f"{prefix} {created} subscribers. "
                f"Skipped {skipped} duplicates. "
                f"{invalid} invalid emails."
            )
        )
