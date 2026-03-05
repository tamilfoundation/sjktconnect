"""Verify school GPS pins against Google Places API locations.

Queries Google Places for each school, compares coordinates, and outputs
an Excel file with variance in metres. Location columns are clickable
Google Maps links.

Usage:
    python manage.py verify_school_pins                          # All schools
    python manage.py verify_school_pins --state Perak            # Filter by state
    python manage.py verify_school_pins --limit 10               # Sample of 10
    python manage.py verify_school_pins --output data/pins.xlsx  # Custom output path
"""

import math
import os
import time

import requests
from django.core.management.base import BaseCommand

from schools.models import School

PLACES_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"


def _get_api_key() -> str:
    return os.environ.get(
        "GOOGLE_MAPS_API_KEY",
        os.environ.get("NEXT_PUBLIC_GOOGLE_MAPS_API_KEY", ""),
    )


def haversine_metres(lat1, lng1, lat2, lng2):
    """Calculate distance in metres between two GPS points."""
    R = 6_371_000  # Earth radius in metres
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def google_maps_link(lat, lng):
    """Return a Google Maps URL for the given coordinates."""
    return f"https://www.google.com/maps?q={lat},{lng}"


class Command(BaseCommand):
    help = "Verify school GPS pins against Google Places API"

    def add_arguments(self, parser):
        parser.add_argument(
            "--state",
            help="Filter schools by state",
        )
        parser.add_argument(
            "--limit",
            type=int,
            help="Limit number of schools to check",
        )
        parser.add_argument(
            "--output",
            default="data/school_pin_verification.xlsx",
            help="Output Excel file path (default: data/school_pin_verification.xlsx)",
        )
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Update database with Google coordinates (skips duplicates and mismatches)",
        )
        parser.add_argument(
            "--skip",
            nargs="*",
            default=[],
            help="MOE codes to skip from auto-update (e.g. --skip CBD7093 JBD2039)",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write("Install openpyxl: pip install openpyxl")
            return

        api_key = _get_api_key()
        if not api_key:
            self.stderr.write("No GOOGLE_MAPS_API_KEY or NEXT_PUBLIC_GOOGLE_MAPS_API_KEY set")
            return

        schools = School.objects.filter(
            gps_lat__isnull=False, gps_lng__isnull=False
        ).order_by("moe_code")

        if options["state"]:
            schools = schools.filter(state__iexact=options["state"])

        if options["limit"]:
            schools = schools[: options["limit"]]

        total = schools.count() if hasattr(schools, "count") else len(schools)
        self.stdout.write(f"Verifying {total} schools against Google Places...\n")

        # Find duplicate school names — these need manual review
        from django.db.models import Count

        dupe_qs = (
            School.objects.values("name")
            .annotate(c=Count("moe_code"))
            .filter(c__gt=1)
        )
        dupe_names = set(d["name"] for d in dupe_qs)
        self.stdout.write(
            f"Found {len(dupe_names)} duplicate names ({sum(d['c'] for d in dupe_qs)} schools) — will flag for manual review\n"
        )

        rows = []
        errors = 0
        applied = 0
        skip_codes = set(options.get("skip") or [])

        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": (
                "places.displayName,places.formattedAddress,"
                "places.location"
            ),
        }

        for i, school in enumerate(schools, 1):
            base_name = school.short_name or school.name
            # Include city/state for disambiguation
            parts = [base_name]
            if school.city:
                parts.append(school.city)
            if school.state:
                parts.append(school.state)
            search_query = ", ".join(parts)
            body = {
                "textQuery": search_query,
                "maxResultCount": 1,
            }
            if school.gps_lat and school.gps_lng:
                body["locationBias"] = {
                    "circle": {
                        "center": {
                            "latitude": float(school.gps_lat),
                            "longitude": float(school.gps_lng),
                        },
                        "radius": 10000.0,
                    }
                }

            try:
                resp = requests.post(
                    PLACES_SEARCH_URL, headers=headers, json=body, timeout=10
                )
                resp.raise_for_status()
                data = resp.json()
            except requests.RequestException as e:
                self.stderr.write(f"  [{i}/{total}] ERROR {school.moe_code}: {e}")
                errors += 1
                continue

            places = data.get("places", [])
            if not places:
                self.stderr.write(
                    f"  [{i}/{total}] NO RESULT {school.moe_code} {school.name}"
                )
                rows.append({
                    "moe_code": school.moe_code,
                    "school_name": school.name,
                    "school_address": school.address,
                    "google_name": "",
                    "google_address": "",
                    "school_loc": google_maps_link(school.gps_lat, school.gps_lng),
                    "google_loc": "",
                    "variance_in_meters": "",
                    "auto_update": "NO — no Google result",
                })
                continue

            place = places[0]
            g_name = place.get("displayName", {}).get("text", "")
            g_address = place.get("formattedAddress", "")
            g_loc = place.get("location", {})
            g_lat = g_loc.get("latitude")
            g_lng = g_loc.get("longitude")

            if g_lat is not None and g_lng is not None:
                variance = round(
                    haversine_metres(
                        float(school.gps_lat), float(school.gps_lng),
                        g_lat, g_lng,
                    )
                )
                g_link = google_maps_link(g_lat, g_lng)
            else:
                variance = ""
                g_link = ""

            s_link = google_maps_link(school.gps_lat, school.gps_lng)

            # Determine if safe to auto-update
            is_duplicate = school.name in dupe_names
            is_skipped = school.moe_code in skip_codes
            # Name match: check if Google's name contains key part of our school name
            our_short = (school.short_name or school.name).upper()
            name_match = (
                our_short in g_name.upper()
                or g_name.upper() in (school.name.upper())
            )

            if is_skipped:
                auto = "SKIP — manual override"
            elif is_duplicate:
                auto = "MANUAL — duplicate name"
            elif not name_match and isinstance(variance, int) and variance > 500:
                auto = "SKIP — name mismatch"
            else:
                auto = "YES"

            # Apply update to database if requested
            if options.get("apply") and auto == "YES" and g_lat is not None:
                school.gps_lat = g_lat
                school.gps_lng = g_lng
                school.save(update_fields=["gps_lat", "gps_lng"])
                applied += 1

            rows.append({
                "moe_code": school.moe_code,
                "school_name": school.name,
                "school_address": school.address,
                "google_name": g_name,
                "google_address": g_address,
                "school_loc": s_link,
                "google_loc": g_link,
                "variance_in_meters": variance,
                "auto_update": auto,
            })

            flag = " ***" if isinstance(variance, int) and variance > 500 else ""
            self.stdout.write(
                f"  [{i}/{total}] {school.moe_code} {school.name}: "
                f"{variance}m{flag}"
            )

            # Rate limit: ~5 QPS
            time.sleep(0.2)

        # Write Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pin Verification"

        col_headers = [
            "moe_code", "school_name", "school_address", "google_name",
            "google_address", "school_loc", "google_loc",
            "variance_in_meters", "auto_update",
        ]
        ws.append(col_headers)

        # Style header row
        from openpyxl.styles import Font, PatternFill

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in rows:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=row["moe_code"])
            ws.cell(row=r, column=2, value=row["school_name"])
            ws.cell(row=r, column=3, value=row["school_address"])
            ws.cell(row=r, column=4, value=row["google_name"])
            ws.cell(row=r, column=5, value=row["google_address"])

            # School location as clickable link
            s_cell = ws.cell(row=r, column=6)
            if row["school_loc"]:
                s_cell.value = row["school_loc"]
                s_cell.hyperlink = row["school_loc"]
                s_cell.font = Font(color="0563C1", underline="single")

            # Google location as clickable link
            g_cell = ws.cell(row=r, column=7)
            if row["google_loc"]:
                g_cell.value = row["google_loc"]
                g_cell.hyperlink = row["google_loc"]
                g_cell.font = Font(color="0563C1", underline="single")

            v_cell = ws.cell(row=r, column=8, value=row["variance_in_meters"])
            # Highlight high variance in red
            if isinstance(row["variance_in_meters"], int) and row["variance_in_meters"] > 500:
                v_cell.font = Font(color="FF0000", bold=True)

            # Auto-update column
            a_cell = ws.cell(row=r, column=9, value=row["auto_update"])
            if row["auto_update"] == "YES":
                a_cell.font = Font(color="008000", bold=True)
            elif row["auto_update"].startswith("MANUAL"):
                a_cell.font = Font(color="FF8C00", bold=True)

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        output_path = options["output"]
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)

        flagged = sum(
            1 for r in rows
            if isinstance(r["variance_in_meters"], int) and r["variance_in_meters"] > 500
        )
        skipped = sum(1 for r in rows if r["auto_update"].startswith("SKIP") or r["auto_update"].startswith("MANUAL"))
        msg = (
            f"\nDone: {len(rows)} schools checked, {flagged} flagged (>500m), "
            f"{skipped} skipped, {errors} errors. Saved to {output_path}"
        )
        if options.get("apply"):
            msg += f"\nApplied Google coordinates to {applied} schools."
        self.stdout.write(self.style.SUCCESS(msg))
