"""
Import SJK(T) schools from MOE Excel file.

Usage:
    python manage.py import_schools ../data/SenaraiSekolahWeb_Januari2026.xlsx
    python manage.py import_schools --dry-run ../data/SenaraiSekolahWeb_Januari2026.xlsx
    python manage.py import_schools --gps-file ../data/school_pin_verification.csv <excel_file>
"""

import csv
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

from openpyxl import load_workbook

from schools.models import Constituency, DUN, School
from schools.utils import to_proper_case, format_phone

# MOE name prefix to replace with short form
MOE_PREFIX = "SEKOLAH JENIS KEBANGSAAN (TAMIL)"
SHORT_PREFIX = "SJK(T)"

# Default GPS verification file
DEFAULT_GPS = Path(__file__).resolve().parent.parent.parent.parent.parent / "data" / "school_pin_verification.csv"


def make_short_name(full_name):
    """Convert MOE full name to short form with proper case.

    'SEKOLAH JENIS KEBANGSAAN (TAMIL) LADANG BIKAM' -> 'SJK(T) Ldg Bikam'
    """
    if MOE_PREFIX in full_name.upper():
        suffix = full_name[len(MOE_PREFIX):].strip()
        return f"{SHORT_PREFIX} {to_proper_case(suffix)}"
    return to_proper_case(full_name)


def parse_code_name(value):
    """Parse 'P140 SEGAMAT' into (code, name). Returns (None, None) if unparseable."""
    if not value or not value.strip():
        return None, None
    value = value.strip()
    match = re.match(r"^([A-Z]\d+)\s+(.+)$", value)
    if match:
        return match.group(1), match.group(2).strip()
    return None, value


def safe_decimal(value):
    """Convert value to Decimal, returning None on failure."""
    if value is None:
        return None
    try:
        d = Decimal(str(value))
        if d == 0:
            return None
        return d
    except (InvalidOperation, ValueError):
        return None


def safe_int(value, default=0):
    """Convert value to int, returning default on failure."""
    if value is None:
        return default
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return default


def load_gps_overrides(gps_file):
    """Load GPS verification CSV and return dict of {moe_code: (lat, lng, verified)}."""
    overrides = {}
    if not gps_file or not Path(gps_file).exists():
        return overrides

    with open(gps_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row.get("school_code", "").strip()
            status = row.get("status", "").strip().lower()
            if code and status == "confirmed":
                lat = safe_decimal(row.get("google_lat") or row.get("lat"))
                lng = safe_decimal(row.get("google_lng") or row.get("lng"))
                if lat and lng:
                    overrides[code] = (lat, lng, True)
    return overrides


class Command(BaseCommand):
    help = "Import SJK(T) schools from MOE Excel file."

    def add_arguments(self, parser):
        parser.add_argument(
            "file",
            type=str,
            help="Path to MOE Excel file (SenaraiSekolahWeb_Januari2026.xlsx)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without creating any records.",
        )
        parser.add_argument(
            "--gps-file",
            type=str,
            default=str(DEFAULT_GPS),
            help=f"Path to GPS verification CSV (default: {DEFAULT_GPS})",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        dry_run = options["dry_run"]
        gps_file = options["gps_file"]

        if not file_path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — no records will be created.\n"))

        # Load GPS overrides
        gps_overrides = load_gps_overrides(gps_file)
        self.stdout.write(f"Loaded {len(gps_overrides)} GPS overrides from verification CSV")

        # Pre-load constituency and DUN lookups
        # MOE Excel has names only (no codes), so build name-based maps
        # Constituency: keyed by uppercase name (MOE uses uppercase)
        # DUN: keyed by (uppercase_dun_name, constituency_pk) since names
        #      can repeat across states
        constituency_by_code = {c.code: c for c in Constituency.objects.all()}
        constituency_by_name = {c.name.upper(): c for c in Constituency.objects.all()}
        dun_by_name = {}
        for d in DUN.objects.select_related("constituency").all():
            dun_by_name[(d.name.upper(), d.constituency_id)] = d
        self.stdout.write(
            f"Loaded {len(constituency_by_code)} constituencies and "
            f"{len(dun_by_name)} DUNs from database"
        )

        # Load Excel
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Read headers from first row
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {h: i for i, h in enumerate(headers) if h}

        stats = {
            "schools_created": 0,
            "schools_updated": 0,
            "skipped_non_sjkt": 0,
            "gps_overridden": 0,
            "constituency_linked": 0,
            "dun_linked": 0,
            "errors": 0,
        }

        def get_cell(row, col_name):
            idx = header_map.get(col_name)
            if idx is None:
                return None
            return row[idx].value if idx < len(row) else None

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                # Filter for SJK(T) only
                jenis = str(get_cell(row, "JENIS/LABEL") or "").strip()
                if "SJK(T)" not in jenis.upper() and "SJKT" not in jenis.upper():
                    stats["skipped_non_sjkt"] += 1
                    continue

                moe_code = str(get_cell(row, "KODSEKOLAH") or "").strip()
                if not moe_code:
                    self.stderr.write(f"  Row {row_num}: No school code — skipped")
                    stats["errors"] += 1
                    continue

                full_name = str(get_cell(row, "NAMASEKOLAH") or "").strip()
                short_name = make_short_name(full_name)

                # GPS: use verification override if available, else MOE data
                # NOTE: MOE columns are swapped: KOORDINATXX = longitude, KOORDINATYY = latitude
                moe_lat = safe_decimal(get_cell(row, "KOORDINATYY"))
                moe_lng = safe_decimal(get_cell(row, "KOORDINATXX"))
                gps_verified = False

                if moe_code in gps_overrides:
                    gps_lat, gps_lng, gps_verified = gps_overrides[moe_code]
                    stats["gps_overridden"] += 1
                else:
                    gps_lat = moe_lat
                    gps_lng = moe_lng

                # Parse constituency and DUN from MOE data
                # MOE format: just name (e.g. " SRI GADING"), or "P140 SEGAMAT"
                parl_raw = str(get_cell(row, "PARLIMEN") or "").strip()
                dun_raw = str(get_cell(row, "DUN") or "").strip()

                # Try code-based lookup first, fall back to name-based
                parl_code, parl_name = parse_code_name(parl_raw)
                dun_code, dun_name = parse_code_name(dun_raw)

                if parl_code:
                    constituency = constituency_by_code.get(parl_code)
                else:
                    constituency = constituency_by_name.get(parl_raw.upper())

                if constituency and dun_code:
                    dun_obj = dun_by_name.get((dun_name.upper() if dun_name else dun_raw.upper(), constituency.code))
                elif constituency:
                    dun_obj = dun_by_name.get((dun_raw.upper(), constituency.code))
                else:
                    dun_obj = None

                if constituency:
                    stats["constituency_linked"] += 1
                if dun_obj:
                    stats["dun_linked"] += 1

                # Also update constituency/DUN state from MOE data
                state = str(get_cell(row, "NEGERI") or "").strip()
                if constituency and not constituency.state and state:
                    constituency.state = state
                    if not dry_run:
                        constituency.save(update_fields=["state"])
                if dun_obj and not dun_obj.state and state:
                    dun_obj.state = state
                    if not dry_run:
                        dun_obj.save(update_fields=["state"])

                # SKM eligibility
                skm_raw = get_cell(row, "SKM<=150")
                skm_eligible = bool(skm_raw and str(skm_raw).strip())

                school_data = {
                    "name": to_proper_case(full_name),
                    "short_name": short_name,
                    "address": to_proper_case(str(get_cell(row, "ALAMATSURAT") or "").strip()),
                    "postcode": str(get_cell(row, "POSKODSURAT") or "").strip(),
                    "city": to_proper_case(str(get_cell(row, "BANDARSURAT") or "").strip()),
                    "state": to_proper_case(state),
                    "ppd": to_proper_case(str(get_cell(row, "PPD") or "").strip()),
                    "constituency": constituency,
                    "dun": dun_obj,
                    "email": str(get_cell(row, "EMAIL") or "").strip().lower(),
                    "phone": format_phone(str(get_cell(row, "NOTELEFON") or "").strip()),
                    "fax": format_phone(str(get_cell(row, "NOFAX") or "").strip()),
                    "gps_lat": gps_lat,
                    "gps_lng": gps_lng,
                    "gps_verified": gps_verified,
                    "enrolment": safe_int(get_cell(row, "ENROLMEN")),
                    "preschool_enrolment": safe_int(get_cell(row, "ENROLMEN PRASEKOLAH")),
                    "special_enrolment": safe_int(get_cell(row, "ENROLMEN KHAS")),
                    "teacher_count": safe_int(get_cell(row, "GURU")),
                    "grade": str(get_cell(row, "GRED") or "").strip(),
                    "assistance_type": str(get_cell(row, "BANTUAN") or "").strip(),
                    "session_count": safe_int(get_cell(row, "BILSESI"), default=1),
                    "session_type": str(get_cell(row, "SESI") or "").strip(),
                    "skm_eligible": skm_eligible,
                    "location_type": str(get_cell(row, "LOKASI") or "").strip(),
                }

                if not dry_run:
                    _, created = School.objects.update_or_create(
                        moe_code=moe_code,
                        defaults=school_data,
                    )
                else:
                    created = not School.objects.filter(moe_code=moe_code).exists()

                if created:
                    stats["schools_created"] += 1
                else:
                    stats["schools_updated"] += 1

            if dry_run:
                transaction.set_rollback(True)

        wb.close()

        # Summary
        self.stdout.write("\n" + "=" * 50)
        self.stdout.write(f"Rows skipped (non-SJK(T)):  {stats['skipped_non_sjkt']}")
        self.stdout.write(f"Schools created:            {stats['schools_created']}")
        self.stdout.write(f"Schools updated:            {stats['schools_updated']}")
        self.stdout.write(f"GPS overridden from CSV:    {stats['gps_overridden']}")
        self.stdout.write(f"Constituency linked:        {stats['constituency_linked']}")
        self.stdout.write(f"DUN linked:                 {stats['dun_linked']}")
        self.stdout.write(f"Errors:                     {stats['errors']}")

        if dry_run:
            self.stdout.write(self.style.WARNING("\nDRY RUN complete — nothing was saved."))
        else:
            self.stdout.write(self.style.SUCCESS(f"\nImport complete. {School.objects.count()} schools in database."))
