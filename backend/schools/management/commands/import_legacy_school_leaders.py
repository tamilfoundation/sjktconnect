"""
Django management command: Import legacy school leadership data from Tamil Foundation 2018

This command imports leadership data from Tamil Foundation's legacy database into SJK(T) Connect.

Features:
  - Matches schools by MOE code (exact match only)
  - Only creates leaders for empty role slots (never overwrites existing)
  - Formats phone to: +60-12 553 7375
  - Formats names to: Title Case with proper titles (Mr., Dr., Dato', etc.)
  - Marks all imported records with data_source="TF_2018"

Usage:
  # Pilot test: 8 representative schools
  python manage.py import_legacy_school_leaders --pilot --dry-run
  python manage.py import_legacy_school_leaders --pilot

  # Full import: all 529 schools
  python manage.py import_legacy_school_leaders --dry-run
  python manage.py import_legacy_school_leaders
"""

import openpyxl
from pathlib import Path
from datetime import date
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from schools.models import School, SchoolLeader


# ============================================================================
# CONFIGURATION
# ============================================================================

LEGACY_DATA_PATH = Path("C:/Users/tamil/Downloads/tamilschool/")

# Pilot test: 8 schools (skip PJS 1 and TROLAK which have no data)
PILOT_MOE_CODES = {
    "BBD8451",  # SJK(T) CASTLEFIELD
    "BBD4055",  # SJK(T) KAJANG
    "ABD6107",  # SJK(T) LADANG HOLYROOD
    "ABD2164",  # SJK(T) CHETTIARS
    "CBD0040",  # SJK(T) LADANG RENJOK
    "NBD1067",  # SJK(T) LADANG JUASSEH
    "MBD1051",  # SJK(T) BATANG MELAKA
    "KBD0058",  # SJK(T) LADANG KUPANG
}

# Role mapping: Legacy column index -> SchoolLeader role
ROLE_MAPPING = {
    35: "headmaster",     # Column 35 = த.ஆ
    39: "board_chair",    # Column 39 = வாரியத் தலைவர்
    36: "pta_chair",      # Column 36 = பெ.ஆ.ச. தலைவர்
    37: "alumni_chair",   # Column 37 = மு.மா. தலைவர்
}

ROLE_DISPLAY = {
    "headmaster": "Headmaster",
    "board_chair": "Board Chairman",
    "pta_chair": "PTA Chairman",
    "alumni_chair": "Alumni Chairman",
}


# ============================================================================
# PHONE & NAME FORMATTING
# ============================================================================

def normalize_phone(phone_str):
    """Convert phone to +60-12 553 7375 format."""
    if not phone_str:
        return ""

    # Remove all non-digits
    digits = ''.join(c for c in str(phone_str) if c.isdigit())

    if not digits:
        return ""

    # Handle Malaysian numbers
    if digits.startswith('60'):
        # Already has country code
        pass
    elif digits.startswith('0'):
        # Local format, convert to international
        digits = '60' + digits[1:]
    else:
        # Assume local format missing leading 0
        digits = '60' + digits

    # Format as +60-XX XXX XXXX
    if len(digits) >= 11 and digits.startswith('60'):
        return f"+{digits[0:2]}-{digits[2:4]} {digits[4:7]} {digits[7:11]}"

    return ""


def normalize_name(name_str):
    """Convert name to Title Case with proper title formatting."""
    if not name_str:
        return ""

    name_str = str(name_str).strip()

    # Common titles to handle
    titles = [
        'datuk', 'dato', 'datin', 'sir', 'madam', 'professor', 'prof',
        'dr', 'mr', 'ms', 'mrs', 'mdm', 'miss', 'en', 'puan',
    ]

    # Split into words
    words = name_str.split()
    if not words:
        return ""

    # Check if first word is a title
    first_lower = words[0].lower().rstrip('.')

    if first_lower in titles:
        title = words[0].lower()
        # Add period if not present
        if not title.endswith('.'):
            title = title + '.'
        # Title case the title
        title = title[0].upper() + title[1:]

        # Title case the rest of the name
        rest = ' '.join(words[1:])
        rest_titlecase = rest.title()

        return f"{title} {rest_titlecase}"
    else:
        # No title, just title case everything
        return name_str.title()


# ============================================================================
# DATA EXTRACTION
# ============================================================================

class Command(BaseCommand):
    help = "Import legacy school leadership data from Tamil Foundation 2018"

    def add_arguments(self, parser):
        parser.add_argument(
            '--pilot',
            action='store_true',
            help='Import only 8 pilot schools for testing',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview what would be imported without saving',
        )

    def handle(self, *args, **options):
        pilot = options['pilot']
        dry_run = options['dry_run']

        self.stdout.write("=" * 100)
        self.stdout.write(
            "IMPORTING LEGACY SCHOOL LEADERSHIP"
            + (" - PILOT (8 Schools)" if pilot else " - FULL (529 Schools)")
            + (" (DRY RUN)" if dry_run else "")
        )
        self.stdout.write("=" * 100)

        # Load staff lookups
        self.stdout.write("\nLoading staff datasets...")
        headmasters = self._load_staff_dataset(
            LEGACY_DATA_PATH / "த.ஆசிரியர் - Status.xlsx",
            "த.ஆசிரியர்"
        )
        board_chairs = self._load_staff_dataset(
            LEGACY_DATA_PATH / "தலைவர் - Status.xlsx",
            "தலைவர்"
        )
        pta_chairs = self._load_staff_dataset(
            LEGACY_DATA_PATH / "பெ.ஆ.ச. - Status.xlsx",
            "பெ.ஆ.ச."
        )
        alumni_chairs = self._load_staff_dataset(
            LEGACY_DATA_PATH / "மு.மாணவர் - Status.xlsx",
            "மு.மாணவர்"
        )

        staff_lookups = {
            "headmaster": headmasters,
            "board_chair": board_chairs,
            "pta_chair": pta_chairs,
            "alumni_chair": alumni_chairs,
        }

        # Load schools dataset
        self.stdout.write("Loading schools dataset...")
        wb = openpyxl.load_workbook(
            LEGACY_DATA_PATH / "பள்ளிகள் - Last view used.xlsx",
            data_only=True
        )
        ws = wb["பள்ளிகள்"]

        legacy_data = {}

        # Extract data for all or pilot schools
        for row_idx in range(2, ws.max_row + 1):
            moe_code = ws.cell(row_idx, 5).value  # Column 5
            school_name = ws.cell(row_idx, 4).value  # Column 4

            if pilot and moe_code not in PILOT_MOE_CODES:
                continue
            if not moe_code:
                continue

            legacy_data[moe_code] = {
                'school_name': school_name,
                'leaders': {}
            }

            # Extract each role
            for col_idx, role in ROLE_MAPPING.items():
                leader_name = ws.cell(row_idx, col_idx).value

                if not leader_name:
                    continue

                # Look up contact info
                lookup = staff_lookups[role]
                contact = lookup.get(leader_name, {'phone': '', 'email': ''})

                legacy_data[moe_code]['leaders'][role] = {
                    'name': leader_name,
                    'phone': contact['phone'],
                    'email': contact['email'],
                    'matched': leader_name in lookup,
                }

        wb.close()

        self.stdout.write(f"\nLoaded {len(legacy_data)} schools from legacy data")

        # Perform import
        if dry_run:
            self._simulate_import(legacy_data)
        else:
            self._commit_import(legacy_data)

    def _load_staff_dataset(self, excel_file, sheet_name):
        """Load staff dataset and create name->contact lookup."""
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb[sheet_name]

        lookup = {}
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row_idx, 4).value  # Column 4
            phone = ws.cell(row_idx, 9).value  # Column 9
            email_work = ws.cell(row_idx, 16).value  # Column 16
            email_home = ws.cell(row_idx, 17).value  # Column 17
            email_other = ws.cell(row_idx, 18).value  # Column 18

            email = email_work or email_home or email_other or ""

            if name:
                lookup[name] = {
                    'phone': phone or "",
                    'email': email or "",
                }

        wb.close()
        return lookup

    def _simulate_import(self, legacy_data):
        """Show what would be imported (dry run)."""
        schools_matched = 0
        schools_not_found = 0
        leaders_created = 0
        leaders_skipped_existing = 0

        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("IMPORT SIMULATION (DRY RUN)")
        self.stdout.write("=" * 100)

        for moe_code in sorted(legacy_data.keys()):
            school_data = legacy_data[moe_code]
            school_name = school_data['school_name']
            leaders = school_data['leaders']

            # Check if school exists in SJK(T) Connect
            try:
                school = School.objects.get(moe_code=moe_code)
                schools_matched += 1
            except School.DoesNotExist:
                schools_not_found += 1
                self.stdout.write(f"\n[NOT FOUND] {school_name} ({moe_code})")
                continue

            if not leaders:
                continue

            self.stdout.write(f"\n[OK] {school_name} ({moe_code})")

            for role, leader_info in leaders.items():
                # Format name and phone
                formatted_name = normalize_name(leader_info['name'])
                formatted_phone = normalize_phone(leader_info['phone'])

                # Check if school already has active leader for this role
                existing = SchoolLeader.objects.filter(
                    school=school,
                    role=role,
                    is_active=True
                ).first()

                if existing:
                    leaders_skipped_existing += 1
                    self.stdout.write(
                        f"  [SKIP] {ROLE_DISPLAY[role]}: {existing.name} (already exists)"
                    )
                else:
                    leaders_created += 1
                    self.stdout.write(
                        f"  [CREATE] {ROLE_DISPLAY[role]}: {formatted_name}"
                    )
                    self.stdout.write(f"    Phone: {formatted_phone or '(empty)'}")
                    self.stdout.write(f"    Email: {leader_info['email'] or '(empty)'}")

        # Summary
        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("SUMMARY")
        self.stdout.write("=" * 100)
        self.stdout.write(f"\nSchools matched: {schools_matched}")
        self.stdout.write(f"Schools not found: {schools_not_found}")
        self.stdout.write(f"Leadership records to create: {leaders_created}")
        self.stdout.write(f"Leadership records skipped (existing): {leaders_skipped_existing}")
        self.stdout.write(f"\nData source: Tamil Foundation, 2018-01-01")
        self.stdout.write("\n(This is a dry run. No data has been modified.)")
        self.stdout.write("To import, run without --dry-run flag")

    @transaction.atomic
    def _commit_import(self, legacy_data):
        """Actually import the data into the database."""
        schools_matched = 0
        schools_not_found = 0
        leaders_created = 0
        leaders_skipped_existing = 0

        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("IMPORTING LEADERSHIP DATA")
        self.stdout.write("=" * 100)

        for moe_code in sorted(legacy_data.keys()):
            school_data = legacy_data[moe_code]
            school_name = school_data['school_name']
            leaders = school_data['leaders']

            # Check if school exists in SJK(T) Connect
            try:
                school = School.objects.get(moe_code=moe_code)
                schools_matched += 1
            except School.DoesNotExist:
                schools_not_found += 1
                self.stdout.write(f"\n[NOT FOUND] {school_name} ({moe_code})")
                continue

            if not leaders:
                continue

            self.stdout.write(f"\n[OK] {school_name} ({moe_code})")

            for role, leader_info in leaders.items():
                # Format name and phone
                formatted_name = normalize_name(leader_info['name'])
                formatted_phone = normalize_phone(leader_info['phone'])

                # Check if school already has active leader for this role
                existing = SchoolLeader.objects.filter(
                    school=school,
                    role=role,
                    is_active=True
                ).first()

                if existing:
                    leaders_skipped_existing += 1
                    self.stdout.write(
                        f"  [SKIP] {ROLE_DISPLAY[role]}: {existing.name} (already exists)"
                    )
                else:
                    # Create new leader record
                    SchoolLeader.objects.create(
                        school=school,
                        role=role,
                        name=formatted_name,
                        phone=formatted_phone,
                        email=leader_info['email'] or "",
                        is_active=True,
                        data_source="TF_2018",
                        data_source_date=date(2018, 1, 1),
                    )
                    leaders_created += 1
                    self.stdout.write(
                        f"  [CREATE] {ROLE_DISPLAY[role]}: {formatted_name}"
                    )
                    self.stdout.write(f"    Phone: {formatted_phone or '(empty)'}")
                    self.stdout.write(f"    Email: {leader_info['email'] or '(empty)'}")

        # Summary
        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("IMPORT COMPLETE")
        self.stdout.write("=" * 100)
        self.stdout.write(f"\nSchools matched: {schools_matched}")
        self.stdout.write(f"Schools not found: {schools_not_found}")
        self.stdout.write(f"Leadership records created: {leaders_created}")
        self.stdout.write(f"Leadership records skipped (existing): {leaders_skipped_existing}")
        self.stdout.write(f"\nData source: Tamil Foundation, 2018-01-01")
        self.stdout.write("\nAll records have been successfully imported to the database.")
