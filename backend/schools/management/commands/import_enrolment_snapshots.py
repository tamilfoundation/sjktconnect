"""Import historical SJKT student-enrolment snapshots from MOE Risalah files.

Each input file is one snapshot date. The MOE schema has shifted twice
across the 5 known historical files:
  * 2018 — JENIS/LABEL = "Jenis Kebangsaan (T)"; student column = MURID
  * 2020 + 2022 + 2025 — JENIS/LABEL = "SJKT"; student column = MURID
  * 2023 — JENIS/LABEL = "SJKT"; student column = ENROLMEN (renamed)

This command auto-detects the schema per file. Skips rows where the MOE
code doesn't match any School in the DB (silently — those are usually
historical schools that have since closed).

Usage:
    python manage.py import_enrolment_snapshots --file <path.xlsx|.csv> --date YYYY-MM-DD
    python manage.py import_enrolment_snapshots --json-bundle <path.json>   # batch import all snapshots

JSON bundle shape:
    {
      "2018-01-31": {"PBD1082": 89, "BBD3056": 23, ...},
      "2020-01-31": {...},
      ...
    }
"""
import csv
import json
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from schools.models import School, SchoolEnrolmentSnapshot

SJKT_LABELS = {"SJKT", "Jenis Kebangsaan (T)", "Jenis Kebangsaan Tamil"}


class Command(BaseCommand):
    help = "Import per-school SJKT enrolment snapshots from MOE Risalah files."

    def add_arguments(self, parser):
        group = parser.add_mutually_exclusive_group(required=True)
        group.add_argument("--file", type=str, help="Single snapshot file (.xlsx or .csv).")
        group.add_argument("--json-bundle", type=str, help="JSON of {date: {moe: students}} for batch import.")
        parser.add_argument("--date", type=str, help="Snapshot date YYYY-MM-DD (required with --file).")
        parser.add_argument("--source", type=str, default="", help="Free-text origin label.")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        if options["file"]:
            if not options["date"]:
                raise CommandError("--date YYYY-MM-DD is required with --file")
            snap_date = datetime.strptime(options["date"], "%Y-%m-%d").date()
            data = self._extract(Path(options["file"]))
            source = options["source"] or Path(options["file"]).name
            self._apply({snap_date: data}, source, options["dry_run"])
        else:
            with open(options["json_bundle"], encoding="utf-8") as f:
                bundle = json.load(f)
            snapshots = {datetime.strptime(d, "%Y-%m-%d").date(): m for d, m in bundle.items()}
            self._apply(snapshots, options["source"], options["dry_run"])

    def _extract(self, path: Path) -> dict:
        """Pull {moe_code: students} from a single MOE Risalah file."""
        if path.suffix.lower() == ".csv":
            return self._extract_csv(path)
        return self._extract_xlsx(path)

    def _extract_csv(self, path: Path) -> dict:
        out = {}
        with open(path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("JENIS/LABEL") in SJKT_LABELS:
                    try:
                        out[row["KODSEKOLAH"]] = int(row.get("MURID") or row.get("ENROLMEN") or 0)
                    except (TypeError, ValueError):
                        pass
        return out

    def _extract_xlsx(self, path: Path) -> dict:
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl required. pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        try:
            ws.calculate_dimension(force=True)
        except Exception:
            pass
        rows = list(ws.iter_rows(values_only=True))
        hdr_idx = next((i for i, r in enumerate(rows) if r and "KODSEKOLAH" in r), None)
        if hdr_idx is None:
            raise CommandError(f"No header row with KODSEKOLAH found in {path.name}")
        hdr = rows[hdr_idx]
        col = {name: i for i, name in enumerate(hdr) if name}
        kod_i = col["KODSEKOLAH"]
        murid_i = col.get("MURID", col.get("ENROLMEN"))
        jenis_i = col["JENIS/LABEL"]
        out = {}
        for r in rows[hdr_idx + 1:]:
            if not r:
                continue
            try:
                if r[jenis_i] in SJKT_LABELS:
                    val = r[murid_i]
                    out[r[kod_i]] = int(val) if val else 0
            except (TypeError, IndexError, ValueError):
                pass
        wb.close()
        return out

    def _apply(self, snapshots: dict, source: str, dry_run: bool):
        moe_to_pk = dict(School.objects.values_list("moe_code", "pk"))
        for snap_date, data in sorted(snapshots.items()):
            matched = 0
            skipped = 0
            for moe, students in data.items():
                pk = moe_to_pk.get(moe)
                if not pk:
                    skipped += 1
                    continue
                matched += 1
                if not dry_run:
                    SchoolEnrolmentSnapshot.objects.update_or_create(
                        school_id=pk,
                        snapshot_date=snap_date,
                        defaults={"students": students, "source": source},
                    )
            verb = "[DRY] would write" if dry_run else "wrote"
            self.stdout.write(
                f"  {snap_date}: {verb} {matched} snapshots, skipped {skipped} (school not in DB)"
            )
