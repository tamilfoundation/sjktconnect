"""Import school bank details from Tamil Foundation Excel database."""

import openpyxl
from django.core.management.base import BaseCommand

from schools.models import School


class Command(BaseCommand):
    help = "Import bank details from TF Excel into School model"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="../data/பள்ளிகள் - மாநிலம்.xlsx",
            help="Path to TF Excel file",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Show what would be imported without saving",
        )

    def handle(self, *args, **options):
        dry_run = options["dry_run"]
        wb = openpyxl.load_workbook(options["file"], read_only=True)
        ws = wb.active

        # Verify column headers match expected positions
        headers = {i: str(v).strip() for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))) if v}
        expected = {4: "குறியீடு", 39: "வாரியக் கணக்கு பெயர்", 40: "கணக்கு எண்", 41: "வங்கி"}
        for col, name in expected.items():
            actual = headers.get(col, "")
            if actual != name:
                self.stderr.write(
                    self.style.ERROR(f"Column {col} expected '{name}', got '{actual}'. Aborting.")
                )
                wb.close()
                return

        updated = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[4]).strip() if row[4] else ""
            acct_name = str(row[39]).strip() if row[39] else ""
            acct_num = str(row[40]).strip() if row[40] else ""
            bank = str(row[41]).strip() if row[41] else ""

            if not code or not acct_num:
                skipped += 1
                continue

            try:
                school = School.objects.get(moe_code=code)
            except School.DoesNotExist:
                self.stdout.write(f"  School not found: {code}")
                skipped += 1
                continue

            if dry_run:
                self.stdout.write(
                    f"  Would update {code}: {bank} / {acct_name} / {acct_num}"
                )
            else:
                school.bank_name = bank
                school.bank_account_name = acct_name
                school.bank_account_number = acct_num
                school.save(
                    update_fields=["bank_name", "bank_account_name", "bank_account_number"]
                )
            updated += 1

        wb.close()

        action = "Would update" if dry_run else "Updated"
        self.stdout.write(
            self.style.SUCCESS(f"{action} {updated} schools, skipped {skipped}")
        )
